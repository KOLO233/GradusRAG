"""Excel 加载器。

使用 openpyxl 解析 .xlsx 文件，将每行数据转为文本。
适合加载结构化数据（如医疗指标表、课程大纲表）。
"""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import List

from src.core.types import Document
from src.ingestion.loaders.base_loader import BaseLoader

logger = logging.getLogger(__name__)


class ExcelLoader(BaseLoader):
    """Excel 文档加载器。

    每个工作表生成一个 Document，表格内容转为 Markdown 表格格式。

    Example:
        >>> loader = ExcelLoader()
        >>> docs = loader.load("data/医疗指标.xlsx")
    """

    def load(self, file_path: str | Path) -> List[Document]:
        file_path = Path(file_path)
        self._validate_file(file_path)

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required. Install: pip install openpyxl"
            )

        logger.info(f"Loading Excel: {file_path.name}")

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        file_hash = self._compute_file_hash(file_path)
        filename = file_path.name

        documents: List[Document] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 转为 Markdown 表格
            text = self._rows_to_markdown(rows, sheet_name)
            if not text.strip():
                continue

            documents.append(Document(
                id=f"{file_hash}_{sheet_name}",
                text=text,
                metadata={
                    "source_path": str(file_path),
                    "filename": filename,
                    "doc_type": "excel",
                    "doc_id": file_hash,
                    "page": len(documents) + 1,
                    "title": sheet_name,
                },
            ))

        wb.close()
        logger.info(f"  Loaded {len(documents)} sheets from {filename}")
        return documents

    @staticmethod
    def _rows_to_markdown(rows: list, sheet_name: str) -> str:
        """将行数据转为 Markdown 表格。"""
        if not rows:
            return ""

        lines = [f"## {sheet_name}\n"]

        # 第一行作为表头
        header = rows[0]
        if header:
            headers = [str(c) if c else "" for c in header]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        # 数据行
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # 跳过全空行
            if all(not c.strip() for c in cells):
                continue
            lines.append("| " + " | ".join(cells) + " |")

        return "\n".join(lines)

    @staticmethod
    def _compute_file_hash(file_path: Path) -> str:
        sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)
        return sha256.hexdigest()[:16]
